import streamlit as st
import openpyxl
from io import BytesIO

st.set_page_config(page_title="Pharma Assay Calculator", layout="wide")

st.title("🧪 HPLC Assay Calculator Portal")
st.info("Enter your data below. All calculations will be preserved in the downloaded Excel file.")

# 1. Input Section
with st.form("assay_form"):
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("General Info")
        product_name = st.text_input("Product / Material Name", "Progesterone")
        report_date = st.date_input("Date of Reporting")
        
        st.subheader("Standard Details")
        potency_std = st.number_input("Standard Potency (%w/w)", value=99.80, format="%.2f")
        std_weight = st.number_input("Standard Taken Weight (mg)", value=0.0, format="%.4f")
        
    with col2:
        st.subheader("Internal Standard (IS) Details")
        is_name = st.text_input("IS Name", "Testosterone")
        potency_is = st.number_input("IS Potency (%w/w)", value=99.75, format="%.2f")
        is_weight = st.number_input("IS Taken Weight (mg)", value=0.0, format="%.4f")

    st.subheader("Standard Response Data (Areas)")
    # Sample data entry for the first 3 injections
    area_data = st.data_editor([
        {"Sr. No": 1, "Std Response": 0, "IS Response": 0},
        {"Sr. No": 2, "Std Response": 0, "IS Response": 0},
        {"Sr. No": 3, "Std Response": 0, "IS Response": 0},
        {"Sr. No": 4, "Std Response": 0, "IS Response": 0},
        {"Sr. No": 5, "Std Response": 0, "IS Response": 0},
    ], num_rows="dynamic")

    submit = st.form_submit_button("Generate Excel Report")

# 2. Excel Generation Logic
if submit:
    # Load your blank template
    wb = openpyxl.load_workbook("Blank_Sheet.xlsx")
    ws = wb.active # Or wb["Assay"]
    
    # Write General Info (Mapping based on your Excel layout)
    ws["D8"] = product_name
    ws["D7"] = str(report_date)
    
    # Write Standard Weights & Potencies
    ws["D13"] = potency_std
    ws["J13"] = potency_is
    ws["C19"] = std_weight
    ws["C22"] = is_weight
    
    # Write Area Data into the rows (starting at Row 26)
    for i, row in enumerate(area_data):
        ws.cell(row=26+i, column=3).value = row["Std Response"]
        ws.cell(row=26+i, column=4).value = row["IS Response"]

    # Save to memory and provide download
    output = BytesIO()
    wb.save(output)
    processed_data = output.getvalue()
    
    st.success("✅ Calculation Complete!")
    st.download_button(
        label="📥 Download Filled Excel File",
        data=processed_data,
        file_name=f"Assay_Report_{product_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
